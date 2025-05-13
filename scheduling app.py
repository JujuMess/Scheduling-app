import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import random
import calendar
from datetime import datetime, date, timedelta

def load_shift_requirements(file_path, sheet_name="Shifts"):
    """
    Load minimum staffing requirements from the Excel file.
    """
    try:
        shift_requirements = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            usecols=["Starting hour", "Minimum required"]
        )
        shift_requirements.columns = ["starting_hour", "min_required"]
        shift_requirements.set_index("starting_hour", inplace=True)
        shift_requirements.index = shift_requirements.index.astype(str)
        return shift_requirements
    except Exception as e:
        print(f"Error loading shift requirements: {e}")
        return None

# NEXT create a function that will populate the rows, based on (f"{}") from the excel sheet. See how we can import this.

def import_sheet(excel_file_path, sheet_name): #function to load a sheet
    try:
        teamsheet = openpyxl.load_workbook(excel_file_path)
        sheet = teamsheet[sheet_name]
        print(f"{excel_file_path} loaded successfully!")
        return sheet
    except FileNotFoundError:
        print(f"The file {excel_file_path} was not found")
    except KeyError:
        print(f"The sheet {sheet_name} was not found")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
    return None

def load_holiday_calendar(file_path, sheet_name):
    try:
        df_hol = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0) #load the excel sheet

        return df_hol
    except Exception as e:
        print(f"Error loading holiday calendar:{e}")
        return None

def process_holiday_calendar(df_hol):
    try:
        df_long = df_hol.melt(ignore_index=False, var_name="Employee", value_name="Status")
        df_long["on_leave"] = df_long["Status"] == "OFF" # replace OFF wit True
        df_long = df_long.drop(columns=["Status"])
        df_long = df_long.reset_index().rename(columns={"index":"Date"}) #make date a normal column
        print(df_long.head())
        return df_long
    except Exception as e:
        print(f"Error processing holiday calendar:{e}")
        return None

def integrate_holidays_into_schedule(df_schedule, df_long):
    df_schedule = df_schedule.merge(
        df_long.rename(columns={"Date": "Day"}),
        on=["Day", "Employee"],
        how="left"
    )
    df_schedule.loc[df_schedule["on_leave"] == True, "shift_time"] = "AL"
    df_schedule = df_schedule.drop(columns=["on_leave"])
    return df_schedule

#creating the schedule structure to be called before integrate holiday
def create_schedule_structure(df_holidays, employee_list, start_date, end_date):
    date_range = pd.date_range(start=start_date, end=end_date) #generate all dates beween the start and end dates

    df_schedule = pd.DataFrame([(day, emp) for day in date_range for emp in employee_list],columns=["Day", "Employee"]) #new table merging both date and employees
    df_schedule["shift_time"] = None #add a column to the df_schedule variable
    return df_schedule

#Loading employee list
def load_employee_list(file_path):
    try:
        df_team = pd.read_excel(file_path, sheet_name="Team")  # Load only the first column
        # Clean column names
        df_team.columns = df_team.columns.str.strip().str.lower()
        df_team = df_team.rename(columns={
            "agent's name": "name",
            "trained social": "trained_social",
            "trained t2": "trained_t2",
            "trained c&c": "trained_cc"
        })
        df_team = df_team.dropna(how="all")  # Remove empty rows if any
        return df_team
    except Exception as e:
        print(f"Error loading employee list: {e}")
        return None

def merge_df_team_and_df_schedule(df_team, df_schedule):
    df_schedule = df_schedule.merge(
        df_team[["name", "trained_social", "team"]],
        how="left",
        left_on="Employee",
        right_on="name"
    )
    return df_schedule

#extracting the social media trained people
def extract_smt_staff_grouped(file_path):
    try:
        # Load the dataset from the Excel file
        df = pd.read_excel(file_path, sheet_name="Team")
        # ✅ Clean column names
        df.columns = df.columns.str.strip().str.lower()
        df = df.rename(columns={
            "agent's name": "name",
            "trained social": "trained_social",
            "trained t2": "trained_t2"
        })

        # ✅ Filter for SMT-trained staff
        smt_team = df[(df["trained_social"] == True) & (df["trained_t2"] == False)]

        # Group by team
        grouped_smt_team = smt_team.groupby("team")["name"].apply(list).to_dict()

        return smt_team, grouped_smt_team
    except Exception as e:
        print(f"Error extracting SMT staff: {e}")
        return None, None

# Checking if at least 2 SMT staff are available per shift
def check_smt_needed(df_schedule):
    grouped = df_schedule.groupby("shift_time")["trained_social"].sum()
    return grouped >= 2  # Returns True if at least 2 SMT are scheduled

def group_employees_by_language(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name="Team")
        df.columns = df.columns.str.strip().str.lower()
        df = df.rename(columns={"agent's name": "name"})
        language_groups = df.groupby("team")["name"].apply(list).to_dict()
        return language_groups
    except Exception as e:
        print(f"Error grouping employees by language: {e}")
        return None
#pick random staff to be used inside the relevant functions
def pick_random_staff(candidates, exclude=None):
    if exclude:
        candidates = [c for c in candidates if c not in exclude]
        if not candidates:
            return None

    return random.choice(candidates)

def assign_initial_shifts(df_schedule):
    df_schedule["Weekday"] = pd.to_datetime(df_schedule["Day"]).dt.weekday
    # Assign 09:00-18:00 to everyone Monday to Saturday
    df_schedule.loc[(df_schedule["Weekday"] < 5) &
                    (df_schedule["shift_time"] != "AL"), "shift_time"] = "9"
    df_schedule.loc[(df_schedule["Weekday"] > 4) &
                    (df_schedule["shift_time"] != "AL"), "shift_time"] = "RDO"

    return df_schedule

def assign_late_late_shift(grouped_smt_team, shift_requirements):
    #print("Before assigning 15:00 shift:", df_schedule[df_schedule["shift_time"] == "15"])
    late_late_team = []
    used_languages = set()
    late_late_needed = shift_requirements.loc["15", "min_required"]
    # Shuffle language keys so it's randomized each time
    languages = list(grouped_smt_team.keys())
    random.shuffle(languages)

    for language in languages:
        candidates = grouped_smt_team[language]
        
        if not candidates:
            continue
        # Pick one random agent from this language
        selected = pick_random_staff(candidates, exclude=late_late_team)
        if selected:
            late_late_team.append(selected)
            used_languages.add(language)
        if len(late_late_team) == late_late_needed:
            break

    return late_late_team, used_languages

def assign_early_late_shift(grouped_all_team, used_languages, late_late_team, shift_requirements):
    #print("Before assigning 13:00 shift:", df_schedule[df_schedule["shift_time"] == "12"])
    early_late_team = []
    early_late_needed = shift_requirements.loc["13", "min_required"]

    languages = list(grouped_all_team.keys())
    random.shuffle(languages)

    for language in languages:
        # Skip languages already used in 15:00 shift
        if language in used_languages:
            continue

        candidates = grouped_all_team[language]
        if not candidates:
            continue

        available_candidates = [c for c in candidates
                                if c not in late_late_team and c not in early_late_team
                                ]
        if not available_candidates:
            continue

        selected = random.choice(available_candidates)
        early_late_team.append(selected)
        used_languages.add(language)  # Mark this language as used

        if len(early_late_team) == early_late_needed:
            break

    return early_late_team, used_languages

def assign_both_late_shifts(df_schedule, late_late_team, early_late_team):
    for person in late_late_team:
        matches = df_schedule[(df_schedule["Employee"] == person) & (df_schedule["Weekday"] < 5)]
        #print(f"🔍 Matches for {person}: {matches.shape[0]} rows")
        if matches.empty:
            print(f"⚠️ No rows matched for {person}, skipping assignment")
        else:
            for _, row in matches.iterrows():
                day_name = row["Day"].day_name()
                assign_one_agent(df_schedule, day_name, "15", person)
                print(f"Assigning {person} on late late shift")

    for person in early_late_team:
        matches = df_schedule[(df_schedule["Employee"] == person) & (df_schedule["Weekday"] < 5)]
        #print(f"🔍 Matches for {person}: {matches.shape[0]} rows")
        if matches.empty:
            print(f"⚠️ No rows matched for {person}, skipping assignment")
        else:
            for _, row in matches.iterrows():
                day_name = row["Day"].day_name()
                assign_one_agent(df_schedule, day_name, "13", person)
                print(f"Assigning {person} on early late shift")


    return df_schedule

def enforce_minimum_staffing(df_schedule, shift_requirements):
    """
    Ensures at least shift_requirements.loc[shift, 'min_required'] people are assigned
    to each (Day, shift_time) in df_schedule.
    """
    # 1. Count current assignments per day & shift
    counts = (
        df_schedule
        .groupby(["Day", "shift_time"])
        .size()
        .rename("assigned_count")
        .reset_index()
    )

    # 2. Merge in the min_required for each shift_time
    counts = counts.merge(
        shift_requirements["min_required"].rename("min_required"),
        left_on="shift_time",
        right_index=True,
        how="left"
    )
    counts = counts[counts["min_required"].notna()]

    # 3. For any (Day, shift_time) where assigned_count < min_required, log or fix
    for _, row in counts.iterrows():
        day, shift_time = row["Day"], row["shift_time"]
        needed = int(row["min_required"] - row["assigned_count"])
        if needed <= 0:
            continue

        print(f"⚠ {day.date()} {shift_time}: need {needed} more staff")

        for _ in range(needed):
            free_df = df_schedule[
                (df_schedule["Day"] == day) &
                (df_schedule["shift_time"].isin(["RDO", "OFF"]))
                ]
            candidates = free_df["Employee"].unique().tolist()
            #only free poeple
            selected = pick_random_staff(candidates, exclude=None)
            if not selected:
                print(f"  ❌ No available candidate for {day.date()} {shift_time}")
                break
            assign_one_agent(df_schedule, day.day_name(), shift_time, selected)

    return df_schedule

def assign_needed_smt(df_schedule, day, shift, smt_needed):
 # will need to merge the smt_team and df_schedule in another function
    # checks all persons who are not working at the shift time and are social trained
    available_smt = df_schedule[(df_schedule["trained_social"] == True) & (df_schedule["shift_time"] != shift)]
    #creates a list of languages for the SMT team
    smt_language = df_schedule[(df_schedule["trained_social"] == True) & (df_schedule["shift_time"] == shift)]["team"].tolist()
    if not available_smt.empty:
        smt_assigned = False # keeps loop on until sortrect person assigned
        while not smt_assigned:
            selected_smt = available_smt.sample(n=smt_needed) #select a random
            selected_lang = selected_smt["team"].tolist()
            if not any(lang in smt_language for lang in selected_lang): # makes sure no language gets doubled in SMT assigned

                for _, smt_row in selected_smt.iterrows():
                    assign_one_agent(df_schedule, day, shift, smt_row["Employee"])

                smt_assigned = True #end loop because smt_assigned

    else:
        print("Assigning one SMT failure: No available SMT")
    return df_schedule

def assign_one_agent(df_schedule, day, shift, agent):
    try:
        df_schedule.loc[
            (df_schedule["Employee"] == agent) &
            (df_schedule["Day"].dt.day_name() == day),
            "shift_time"
        ] = shift
    except Exception as e:
        print(f"error {e} when assigning agent")

def assign_saturday_late_shifts(df_schedule, grouped_smt_team, grouped_all_team):

    sat_late_late_team, used_languages = assign_late_late_shift(grouped_smt_team)
    sat_early_late_team, _ = assign_early_late_shift(grouped_all_team, used_languages, sat_late_late_team)

    for employee in sat_late_late_team:
        match = df_schedule[(df_schedule["Employee"] == employee) &
                            (df_schedule["Day"].dt.day_name() == "Saturday")]
        for idx, row in match.iterrows():
            assign_one_agent(df_schedule, row["Day"], "15", employee)
    for employee in sat_early_late_team:
        match = df_schedule[(df_schedule["Employee"] == employee) &
                            (df_schedule["Day"].dt.day_name() == "Saturday")]
        for idx, row in match.iterrows():
            assign_one_agent(df_schedule, row["Day"], "13", employee)

    return df_schedule, sat_late_late_team, sat_early_late_team

def assign_saturdays_and_rdo(df_schedule, df_team, sat_late_late_team, sat_early_late_team):
    sat_morning_team = set()
    languages_morning_team = set()
    morning_shift = "9"
    #assign morning
    morning_smt_needed = 2 #number of time morning SMT will be added

    #asigned SMT to evening and morning
    assign_needed_smt(df_schedule, "Saturday", morning_shift, morning_smt_needed)

    #updates the list of people working in the morning
    morning_names = df_schedule.loc[
        (df_schedule["Day"].dt.day_name() == "Saturday") &
        (df_schedule["shift_time"] == morning_shift),
        "Employee"
    ].tolist()
    sat_morning_team.update(morning_names)
    # updates the list of people working in the evening

    #start a  loop to assign morning team
    while len(sat_morning_team) < 6 :
        for _, row in df_team.sample(frac=1).iterrows(): #iterate through the whole df_team
            if row["team"] in languages_morning_team or row["name"] in sat_morning_team or row["name"] in sat_late_late_team or row["name"] in sat_early_late_team: #skip people whose language is in the morning team
                continue
            assign_one_agent(df_schedule, "Saturday", morning_shift, row["name"]) #use the function to assign an agent
            languages_morning_team.add(row["team"]) # add the language to the morning language list
            sat_morning_team.add(row["name"]) # add agents to the morning team

    all_saturday_agents = set(sat_early_late_team + sat_late_late_team + list(sat_morning_team))
    possible_day_off = ["Wednesday", "Thursday", "Friday"]
    for i, agent in enumerate(all_saturday_agents):
        day_off = possible_day_off[i % len(possible_day_off)]
        df_schedule.loc[
            (df_schedule["Employee"] == agent) &
            (df_schedule["Day"].dt.day_name() == day_off),
         "shift_time"
        ] = "RDO"

    return df_schedule, sat_morning_team, sat_late_late_team, sat_early_late_team

def enforce_max_days_off(df_schedule, df_team):
    for _, row in df_team.iterrows():
        agent = row["name"]
        max_rdo_per_week = 7 - row["days_per_week"]
        agent_schedule = df_schedule[df_schedule["Employee"] == agent].copy()
        agent_schedule["temp_week"] = agent_schedule["Day"].dt.isocalendar().week
        for week, week_data in agent_schedule.groupby("temp_week"):
            rdo_count = week_data[week_data["shift_time"] == "RDO"].shape[0]
            rdo_balance = max_rdo_per_week - rdo_count

            if rdo_balance < 0:
                rdo_rows = week_data[week_data["shift_time"] == "RDO"]
                rows_to_change = rdo_rows.sample(n=abs(rdo_balance)).index
                for idx in rows_to_change:
                    target_day = df_schedule.loc[idx, "Day"]
                    assign_one_agent(df_schedule, target_day.day_name(), "9", agent)

            elif rdo_balance > 0:
                working_rows = week_data[week_data["shift_time"] != "RDO"]
                rows_to_change = working_rows.sample(n=rdo_balance).index
                for idx in rows_to_change:
                    target_day = df_schedule.loc[idx, "Day"]
                    assign_one_agent(df_schedule, target_day.day_name(), "RDO", agent)
    return df_schedule

def smt_check(df_schedule, df_team):
    team_shifts = ["9", "13", "15"]
    smt_team = df_team[df_team["trained_social"] == True]["name"].tolist()
    grouped = df_schedule.groupby(["Day", "shift_time"])

    for (day, shift), group_df in grouped:
        if shift not in team_shifts:
            continue
        scheduled_agents = group_df["Employee"].tolist()
        smts_present = [agent for agent in scheduled_agents if agent in smt_team]

    if len(smts_present) >= 2:
        print(f"✅ Enough SMT on {day} ({shift}): {len(smts_present)}")
    else:
        print(f"❌ Not enough SMT on {day} ({shift}): {len(smts_present)}")

def report_smt_coverage(df_schedule, df_team):

    smt_team = df_team[df_team["trained_social"] == True]["name"].tolist()
    report_df = df_schedule[df_schedule["Employee"].isin(smt_team)]
    pivot = report_df.pivot_table(index = "shift_time", columns = "Day", values="Employee", aggfunc = "count", fill_value = 0)
    print("SMT report created!")
    return pivot

def check_and_fill_smt(df_schedule, shift_requirements, df_team):
    """
    After generating the SMT coverage report, check if any shifts are under-staffed
    and assign more SMT staff if needed.
    """
    smt_team = df_team[df_team["trained_social"] == True]["name"].tolist()
    # Generate the SMT coverage report
    smt_coverage = report_smt_coverage(df_schedule, df_team)

    # Check for each shift if we need more SMT staff
    for (shift_time, day), count in smt_coverage.stack().items():
        if shift_time not in shift_requirements.index:
            continue
        min_required = shift_requirements.loc[shift_time, "min_required"]
        if count < min_required:
            smt_needed = min_required - count
            print(f"⚠ {day} {shift_time}: Need {smt_needed} more SMT staff.")

            # Assign more SMT staff to this shift
            available_smt = df_schedule[(df_schedule["shift_time"] != shift_time) &
                                        (df_schedule["shift_time"] != "AL") &
                                        (df_schedule["shift_time"] != "RDO") &
                                        (df_schedule["Employee"].isin(smt_team))]
            for _ in range(smt_needed):
                selected_smt = random.choice(available_smt["Employee"].tolist())
                assign_one_agent(df_schedule, day, shift_time, selected_smt)
                print(f"{selected_smt} added to {day} at {shift_time}")

    return df_schedule

def output_schedule(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    # 🗓️ Step 1: Get next and previous months
    try:
        # Extract all sheets that look like "Month YYYY"
        schedule_tabs = []
        for name in wb.sheetnames:
            try:
                parsed = datetime.strptime(name, "%B %Y")
                schedule_tabs.append(parsed)
            except ValueError:
                continue  # Skip tabs that don't match

        if not schedule_tabs:
            raise ValueError("No valid schedule tabs found")

        # Get the most recent tab
        prev_month = max(schedule_tabs)

        # Compute next month manually
        if prev_month.month == 12:
            next_month = datetime(prev_month.year + 1, 1, 1)
        else:
            next_month = datetime(prev_month.year, prev_month.month + 1, 1)

    except Exception as e:
        print(f"Error {e}: no previous schedule found, defaulting to current month")
        today = datetime.today()
        if today.month == 12:
            next_month = datetime(today.year + 1, 1, 1)
            prev_month = datetime(today.year, 12, 1)
        else:
            next_month = datetime(today.year, today.month + 1, 1)
            prev_month = datetime(today.year, today.month, 1)

    next_sheet_name = next_month.strftime("%B %Y")
    prev_sheet_name = prev_month.strftime("%B %Y")
    base_sheet_name = "Base schedule"
    #define next year and month
    year = next_month.year
    month = next_month.month

    # 📁 Step 2: Load workbook and create new sheet from base
    base_sheet = wb[base_sheet_name]
    new_sheet = wb.copy_worksheet(base_sheet)
    new_sheet.title = next_sheet_name
    print(f"New tab '{next_sheet_name}' created successfully")

    # 📅 Step 3: Determine start and end date
    try:
      previous_sheet = wb[prev_sheet_name]
      for col in reversed(range(3, previous_sheet.max_column + 1)):
        val = previous_sheet.cell(row=1, column=col).value
        if isinstance(val, (datetime, date)):
          previous_sheet_last_date = val
          break
    except KeyError as e:
      print(f"Error {e}: no previous month tab")
      first_day = datetime(year, month, 1)
      previous_sheet_last_date = first_day - timedelta(days=1)


    start_date = previous_sheet_last_date + timedelta(days=1)
    end_of_month = datetime(year, month, calendar.monthrange(year, month)[1]).date()
    end_date = end_of_month + timedelta(days=(6 - end_of_month.weekday()))

    # 📆 Step 4: Write full range of dates into row 1 (starting column 3)
    for i, current_date in enumerate(pd.date_range(start=start_date, end=end_date)):
        new_sheet.cell(row=1, column=3 + i).value = current_date.date()

    # 📋 Step 5: Fill weekday names in each "team" row
    for row in range(1, new_sheet.max_row + 1):
        team_cell = new_sheet.cell(row=row, column=1)
        if team_cell.value and "team" in str(team_cell.value).strip().lower():
            print(f"Found team row: {row} → {team_cell.value}")
            for col in range(3, new_sheet.max_column + 1):
                date_cell = new_sheet.cell(row=1, column=col).value
                if isinstance(date_cell, (datetime, date)):
                    weekday = date_cell.strftime("%A")
                    new_sheet.cell(row=row, column=col).value = weekday
                else:
                    print(f"Warning: cell at row=1, col={col} is not a valid date.")

    #step 6: Fill the formulas at the bottom of the sheet
    latest_row = new_sheet.max_row

    for col in range(3, new_sheet.max_column + 1):
        for row in range(latest_row - 3, latest_row):
            col_letter = get_column_letter(col)
            table_end = latest_row - 5
            new_sheet.cell(row=row, column=col).value = f"=COUNTIF({col_letter}3:{col_letter}{table_end},B{row})"

        new_sheet.cell(row=latest_row, column=col).value = (f"={col_letter}{latest_row-3}+"
                                                                     f"{col_letter}{latest_row-2}+"
                                                                     f"{col_letter}{latest_row-1}")


    # 💾 Step 6: Save updated file
    wb.save("Team Rota 2024-2025 - Team_updated.xlsx")




####MAIN####
# ---------------------------
# 📁 1. File setup
# ---------------------------
excel_file = "/home/julien/Documents/PycharmProjects/PythonschedulingNEW/Team Rota 2024-2025 - Team.xlsx"


# ▶️ 1.a Load your shift-requirements sheet once
shift_requirements = load_shift_requirements(excel_file, sheet_name="Shifts")
# ---------------------------
# 🧑‍💼 2. Load employee list
# ---------------------------
employee_list = load_employee_list(excel_file)
df_team = pd.read_excel(excel_file, sheet_name="Team")
# 🧹 Clean column names
df_team = df_team.loc[:, ~df_team.columns.str.contains("Unnamed")]  # Drop 'Unnamed' columns
df_team.columns = df_team.columns.str.strip().str.lower()  # Clean and lowercase
df_team = df_team.rename(columns={
    "agent's name": "name",
    "days per week": "days_per_week",
    "hours per week": "hours_per_week",
    "trained social": "trained_social",
    "trained t2": "trained_t2",
    "trained c&c": "trained_cc",  # if you use this one too
})

# ---------------------------
# 🗓️ 3. Create empty schedule
# ---------------------------
df_schedule = create_schedule_structure(
    df_holidays=None,
    employee_list=employee_list["name"].tolist(),
    start_date="2025-04-01",
    end_date="2025-04-30"
)

# 3.5 extract SMT staff
smt_team, grouped_smt_team = extract_smt_staff_grouped(excel_file)
smt_team["shift_time"] = None
# ---------------------------
# 🌐 4. Group employees by language
# ---------------------------
language_groups = group_employees_by_language(excel_file)
#print(df_schedule.columns)
df_schedule = merge_df_team_and_df_schedule(df_team, df_schedule)

# Inspect to confirm that the merge worked:
print(df_schedule.head(10)[['Employee','team','trained_social','shift_time']])
# ---------------------------
# 10. Load and process the holiday calendar
# ---------------------------
df_holidays = load_holiday_calendar(excel_file, sheet_name="Holidays")
df_holidays_long = process_holiday_calendar(df_holidays)
if df_holidays_long is not None:
    df_schedule = integrate_holidays_into_schedule(df_schedule, df_holidays_long)
else:
    print("⚠ Skipping holiday integration—no valid holiday data loaded.")
# ↓ Remove any other bare calls to integrate_holidays_into_schedule below


# ---------------------------
# ⏰ 5. Assign everyone 9:00 shifts
# ---------------------------
df_schedule = assign_initial_shifts(df_schedule)
#print("✅ After assigning initial shifts:")
#print(df_schedule.head(10))

grouped_all_team = df_team.groupby("team")["name"].apply(list).to_dict()


# ---------------------------
# 🌙 7. Assign late shift teams
# ---------------------------
late_late_team, used_languages = assign_late_late_shift(grouped_smt_team, shift_requirements)
early_late_team, used_languages = assign_early_late_shift(grouped_smt_team, used_languages, late_late_team, shift_requirements)

# ---------------------------
# 📝 8. Apply night shifts to the schedule as well as saturday shifts
# ---------------------------
df_schedule = assign_both_late_shifts(df_schedule, late_late_team, early_late_team)
print(df_schedule.columns)

df_schedule, sat_morning_team, sat_late_late_team, sat_early_late_team = assign_saturdays_and_rdo(
    df_schedule,
    df_team,
    late_late_team,
    early_late_team
)


# ---------------------------
#12 Checks all requirements are met
# ---------------------------
enforce_minimum_staffing(df_schedule, shift_requirements)
smt_check(df_schedule, df_team)
report_smt_coverage(df_schedule, df_team)

enforce_max_days_off(df_schedule, df_team)

print("\n–– SMT Coverage Before Fill ––")
print(report_smt_coverage(df_schedule, df_team))
df_schedule = check_and_fill_smt(df_schedule, shift_requirements, df_team)
print("\n–– SMT Coverage After Fill ––")
print(report_smt_coverage(df_schedule, df_team))

output_schedule("/home/julien/Documents/PycharmProjects/PythonschedulingNEW/Team Rota 2024-2025 - Team.xlsx")


"""""
sat_morning_check = df_schedule[
    (df_schedule["Day"].dt.day_name() == "Saturday") &
    (df_schedule["Employee"].isin(sat_morning_team))
]

sat_evening_check = df_schedule[
    (df_schedule["Day"].dt.day_name() == "Saturday") &
    (df_schedule["Employee"].isin(sat_late_late_team | sat_early_late_team))
]




#print("📅 Saturday Morning Assignments:")
#print(sat_morning_check[["Employee", "shift_time"]].drop_duplicates())

#print("📅 Saturday Evening Assignments:")
#print(sat_evening_check[["Employee", "shift_time"]].drop_duplicates())

# Days we're distributing OFFs on
days_off = ["Wednesday", "Thursday", "Friday"]

for team, label in [(sat_morning_team, "Morning"), (saturday_evening_team, "Evening")]:
    print(f"🕵️ Checking OFFs for {label} team:")
    for agent in team:
        off_count = df_schedule[
            (df_schedule["Employee"] == agent) &
            (df_schedule["Day"].dt.day_name().isin(days_off)) &
            (df_schedule["shift_time"] == "RDO")
        ].shape[0]
        print(f"  {agent} ➜ {off_count} weekday OFF(s)")

enforce_max_days_off(df_schedule, df_team)

days_off = ["Wednesday", "Thursday", "Friday"]

for team, label in [(sat_morning_team, "Morning"), (saturday_evening_team, "Evening")]:
    print(f"🕵️ Checking OFFs for {label} team:")
    for agent in team:
        off_count = df_schedule[
            (df_schedule["Employee"] == agent) &
            (df_schedule["Day"].dt.day_name().isin(days_off)) &
            (df_schedule["shift_time"] == "RDO")
        ].shape[0]
        print(f"  {agent} ➜ {off_count} weekday OFF(s)")"""