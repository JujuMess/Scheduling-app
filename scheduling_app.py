import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import os
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import pandas as pd
import random
import numpy as np
import calendar
from datetime import datetime, date, timedelta
from dateutil.parser import parse

def load_shift_requirements(excel_file_path, sheet_name="Shifts"):
    """
    Load minimum staffing requirements from the Excel file.
    """
    try:
        shift_requirements = pd.read_excel(
            excel_file_path,
            sheet_name=sheet_name,
            usecols=["Starting hour", "Minimum required", "Social agents needed"]
        )
        shift_requirements.rename(columns={"Social agents needed": "smt_needed"},
    inplace=True)
        shift_requirements.columns = ["starting_hour", "min_required", "smt_needed"]
        shift_requirements.set_index("starting_hour", inplace=True)
        shift_requirements.index = shift_requirements.index.astype(str)
        return shift_requirements
    except Exception as e:
        print(f"Error loading shift requirements: {e}")
        return None

def import_sheet(excel_file_path, sheet_name): #function to load a sheet
    try:
        teamsheet = openpyxl.load_workbook(excel_file_path)
        sheet = teamsheet[sheet_name]
        print(f"{excel_file_path} loaded successfully!")
        return sheet
    except FileNotFoundError:
        print(f"The file {excel_file_path} was not found")
    except KeyError:
        print(f"The sheet {sheet_name} was not found")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
    return None

def load_holiday_calendar(file_path, sheet_name):
    try:
        df_hol = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0) #load the excel sheet

        return df_hol
    except Exception as e:
        print(f"Error loading holiday calendar:{e}")
        return None

def process_holiday_calendar(df_hol):
    try:
        df_long = df_hol.melt(ignore_index=False, var_name="Employee", value_name="Status")
        df_long["on_leave"] = df_long["Status"] == "OFF" # replace OFF wit True
        df_long = df_long.drop(columns=["Status"])
        df_long = df_long.reset_index().rename(columns={"index":"Date"}) #make date a normal column
        df_long["Employee"] = (
            df_long["Employee"].astype(str)
            .str.replace("\u00A0", " ", regex=False)
            .str.strip()
            .str.replace(r"\*$", "", regex=True)
            .str.lower()
        )

        print(df_long.head())
        return df_long
    except Exception as e:
        print(f"Error processing holiday calendar:{e}")
        return None

def integrate_holidays_into_schedule(df_schedule, df_long):
    df_schedule = df_schedule.merge(
        df_long.rename(columns={"Date": "Day"}),
        on=["Day", "Employee"],
        how="left"
    )
    df_schedule.loc[df_schedule["on_leave"] == True, "shift_time"] = "AL"
    df_schedule = df_schedule.drop(columns=["on_leave"])
    return df_schedule

def _norm_name_list(names):
    return [str(n).strip().replace("*","").lower() for n in names if pd.notna(n)]


#creating the schedule structure to be called before integrate holiday
def create_schedule_structure(df_holidays, employee_list, start_date, end_date):
    date_range = pd.date_range(start=start_date, end=end_date) #generate all dates beween the start and end dates

    df_schedule = pd.DataFrame([(day, emp) for day in date_range for emp in employee_list],columns=["Day", "Employee"]) #new table merging both date and employees
    df_schedule["Employee"] = (
        df_schedule["Employee"].astype(str)
        .str.strip()
        .str.replace(r"\*$", "", regex=True)
        .str.lower()
    )

    df_schedule["shift_time"] = None #add a column to the df_schedule variable
    return df_schedule

#Loading employee list
def normalize_smt_column(series):
    return (
        series
        .apply(lambda x: str(x).strip().lower() if pd.notna(x) else "")
        .map({"yes": True, "true": True, "1": True, "y": True, "✓": True})
        .fillna(False)
    )

def load_employee_list(file_path):
    try:
        df_team = pd.read_excel(file_path, sheet_name="Team")

        # Clean column names
        df_team.columns = df_team.columns.str.strip().str.lower()
        print("🧾 Raw values from 'trained social':", df_team["trained social"].unique())
        df_team = df_team.rename(columns={
            "agent's name": "name",
            "trained social": "trained_social",
            "trained t2": "trained_t2",
            "trained c&c": "trained_cc",
            "days per week": "days_per_week",
            "hours per week": "hours_per_week",
        })

        df_team["norm_name"] = (
            df_team["name"].astype(str)
            .str.replace("\u00A0", " ", regex=False)  # NBSP → space
            .str.strip()
            .str.replace(r"\*$", "", regex=True)  # drop trailing *
            .str.lower()
        )

        name_map = dict(zip(df_team["norm_name"], df_team["name"]))

        # ─── Normalize the TRUE/FALSE strings into real booleans ───
        df_team["trained_social"] = (
            df_team["trained_social"]
              .astype(str)             # ensure strings
              .str.strip()             # trim whitespace
              .str.upper()             # "True" → "TRUE"
              .map({"TRUE": True, "FALSE": False})
              .fillna(False)           # anything else → False
        )

        print("✅ trained_social value counts:", df_team["trained_social"].value_counts())

        df_team = df_team.dropna(how="all")  # Remove empty rows if any
        return df_team, name_map
    except Exception as e:
        print(f"Error loading employee list: {e}")
        return None


def merge_df_team_and_df_schedule(df_team, df_schedule):
    df_schedule = df_schedule.merge(
        df_team[["name", "trained_social", "team"]],
        how="left",
        left_on="Employee",
        right_on="name"
    )
    return df_schedule

#extracting the social media trained people
def extract_smt_staff_grouped(file_path):
    try:
        # Load the dataset from the Excel file
        df = pd.read_excel(file_path, sheet_name="Team")
        # ✅ Clean column names
        df.columns = df.columns.str.strip().str.lower()
        df = df.rename(columns={
            "agent's name": "name",
            "trained social": "trained_social",
            "trained t2": "trained_t2",
            "team": "team"
        })
        df["trained_social"] = normalize_smt_column(df["trained_social"])

        # ✅ Filter for SMT-trained staff
        smt_team = df[(df["trained_social"] == True) & (df["trained_t2"] == False)]

        # Group by team
        grouped_smt_team = smt_team.groupby("team")["name"].apply(list).to_dict()

        return smt_team, grouped_smt_team
    except Exception as e:
        print(f"Error extracting SMT staff: {e}")
        return None, None

# Checking if at least 2 SMT staff are available per shift
def check_smt_needed(df_schedule):
    grouped = df_schedule.groupby("shift_time")["trained_social"].sum()
    return grouped >= 2  # Returns True if at least 2 SMT are scheduled

def group_employees_by_language(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name="Team")
        df.columns = df.columns.str.strip().str.lower()
        df = df.rename(columns={"agent's name": "name"})
        language_groups = df.groupby("team")["name"].apply(list).to_dict()
        return language_groups
    except Exception as e:
        print(f"Error grouping employees by language: {e}")
        return None
#pick random staff to be used inside the relevant functions
def pick_random_staff(candidates, exclude=None):
    # If no candidates at all, return None
    if not candidates:
        return None
    # Exclude any names in the exclude list
    if exclude:
        candidates = [c for c in candidates if c not in exclude]
        if not candidates:
            return None
    # Pick randomly from the remaining candidates
    return random.choice(candidates)

def assign_initial_shifts(df_schedule):
    df_schedule["Weekday"] = pd.to_datetime(df_schedule["Day"]).dt.weekday
    # Assign 09:00-18:00 to everyone Monday to Saturday
    df_schedule.loc[(df_schedule["Weekday"] < 5) &
                    (df_schedule["shift_time"] != "AL"), "shift_time"] = "9"
    df_schedule.loc[(df_schedule["Weekday"] > 4) &
                    (df_schedule["shift_time"] != "AL"), "shift_time"] = "RDO"

    return df_schedule

# at top of file
import numpy as np

def update_history(df_history, df_week, **kwargs):
    """
    kwargs can include any of:
      - weekday_late_late, weekday_early_late
      - saturday_late_late, saturday_early_late, saturday_morning
    Each value should be a list[str] of employee names.
    """

    # Map shift keys → (latest_column, count_column)
    key_map = {
        "weekday_late_late":    ("latest_late_late",            "count_late_late"),
        "weekday_early_late":   ("latest_early_late",           "count_early_late"),
        "saturday_morning":     ("latest_saturday_morning",     "count_saturday_morning"),
        "saturday_early_late":  ("latest_saturday_early_late",  "count_saturday_early_late"),
        "saturday_late_late":   ("latest_saturday_late_late",   "count_saturday_late_late"),
    }
    week_anchor = pd.to_datetime(df_week["Day"].max()) if not df_week.empty else pd.NaT

    # Ensure dtypes are friendly
    for _, (latest_col, count_col) in key_map.items():
        if latest_col in df_history.columns:
            df_history[latest_col] = pd.to_datetime(df_history[latest_col], errors="coerce")
        if count_col in df_history.columns:
            df_history[count_col] = pd.to_numeric(df_history[count_col], errors="coerce").fillna(0).astype(int)

    # For each provided key, update counts and latest
    for key, names in kwargs.items():
        if key not in key_map:
            continue
        latest_col, count_col = key_map[key]
        name_list = _to_name_list(names)
        if not name_list:
            continue

        for name in name_list:
            # increment count
            row_mask = (df_history["name"].astype(str).str.strip().str.lower()
                        == str(name).strip().lower())
            if not row_mask.any():
                continue  # name not in history table

            df_history.loc[row_mask, count_col] = df_history.loc[row_mask, count_col].fillna(0).astype(int) + 1

            # latest = max day in the currently assigned week for that person
            # (no need to re-classify shifts)
            latest_day = df_week.loc[
                df_week["Employee"].astype(str).str.strip().str.lower()
                == str(name).strip().lower(), "Day"
            ].max()

            # Fallback to the week’s last day if no specific match was found
            if pd.isna(latest_day):
                latest_day = week_anchor

            if pd.notna(latest_day):
                df_history.loc[row_mask, latest_col] = pd.to_datetime(latest_day)

            df_history.loc[row_mask, latest_col] = pd.to_datetime(latest_day)

    return df_history

def select_fairest(candidates, df_history, df_team,
                   shift_type, k=1,
                   week_start=None, cooldown_days=6,
                   filter_func=None, duplicate_teams=True):
    """
    Return a list of k norm_names chosen fairly for `shift_type`.
    Uses counts + latest from df_history, optional filter_func, and a week-based RNG tie-breaker.
    """
    import pandas as pd
    import numpy as np

    # --- Normalize candidate names to match df_team.norm_name ---
    cand = (pd.Series(candidates, dtype="object")
              .dropna()
              .astype(str)
              .str.replace("\u00A0", " ", regex=False)
              .str.strip()
              .str.replace(r"\*$", "", regex=True)
              .str.lower())
    candidates_df = pd.DataFrame({"name": cand.unique()})

    # --- Merge team metadata on normalized key ---
    candidates_df = candidates_df.merge(
        df_team[["norm_name", "team", "trained_social"]],
        left_on="name", right_on="norm_name", how="left"
    )

    # --- Optional external filter (row-wise) ---
    if filter_func is not None and len(candidates_df):
        mask = candidates_df.apply(filter_func, axis=1).astype(bool)
        candidates_df = candidates_df[mask]

    # --- Weekly cooldown: build a mask ALIGNED TO df_history.index (no set_index here) ---
    latest_col = f"latest_{shift_type}"
    if week_start is not None and latest_col in df_history.columns and len(candidates_df):
        s = df_history[latest_col]

        # handle object cells that may contain arrays/lists
        if s.dtype == "O":
            s = s.apply(lambda x: (x[0] if isinstance(x, (list, tuple, np.ndarray)) else x))

        s = pd.to_datetime(s, errors="coerce")
        # normalize tz if present
        try:
            s = s.dt.tz_convert(None)
        except Exception:
            try:
                s = s.dt.tz_localize(None)
            except Exception:
                pass

        cutoff = pd.Timestamp(week_start).normalize() - pd.Timedelta(days=cooldown_days)
        mask = s.ge(cutoff)   # NaT → False

        # names to block (normalize like candidates)
        recent_block = set(
            df_history.loc[mask, "name"]
            .astype(str)
            .str.replace("\u00A0", " ", regex=False)
            .str.strip()
            .str.replace(r"\*$", "", regex=True)
            .str.lower()
        )
        candidates_df = candidates_df[~candidates_df["name"].isin(recent_block)]

    if not len(candidates_df):
        return []

    # --- Team duplication control (simple: max 1 per team when False) ---
    if not duplicate_teams and "team" in candidates_df.columns:
        candidates_df = candidates_df.drop_duplicates(subset=["team"])

    # --- Fairness: min count, then oldest latest, then week-seeded random ---
    hist = df_history.copy()
    hist_idx = (hist["name"].astype(str)
                          .str.replace("\u00A0", " ", regex=False)
                          .str.strip()
                          .str.replace(r"\*$", "", regex=True)
                          .str.lower())
    hist.index = hist_idx

    cnt_col = f"count_{shift_type}"
    if cnt_col in hist.columns:
        counts = hist.reindex(candidates_df["name"])[cnt_col].fillna(0)
        candidates_df = candidates_df.assign(_cnt=counts.values)
        min_cnt = candidates_df["_cnt"].min()
        tier = candidates_df[candidates_df["_cnt"] == min_cnt].copy()
    else:
        tier = candidates_df.copy()

    latest_series = pd.to_datetime(
        hist.reindex(tier["name"])[latest_col],
        errors="coerce"
    ).fillna(pd.Timestamp("1970-01-01"))
    tier = tier.assign(_latest=latest_series.values)

    # Oldest (earliest) latest date first; then sample to break remaining ties
    tier = tier.sort_values("_latest", ascending=True)

    n_pick = min(k, len(tier))
    if n_pick <= 0:
        return []

    seed = int(pd.Timestamp(week_start).strftime("%Y%W")) if week_start is not None else None
    picked = tier.sample(n=n_pick, random_state=seed)["name"].tolist()
    return picked


def assign_both_late_shifts(df_week, df_history, df_team, grouped_smt_team, shift_requirements):
    """
    Pick weekday SMTs for 15:00 (late_late) and 13:00 (early_late),
    apply weekly cooldown using week_start,
    then fill remaining headcount with non-SMT agents.
    Returns: df_week, late_late_team (list), early_late_team (list)
    """
    import pandas as pd

    # Week context for cooldown
    week_start = pd.to_datetime(df_week["Day"].min()).normalize()

    # Requirements
    smt_req_13   = int(shift_requirements.loc["13", "smt_needed"])
    smt_req_15   = int(shift_requirements.loc["15", "smt_needed"])
    headcount_13 = int(shift_requirements.loc["13", "min_required"])
    headcount_15 = int(shift_requirements.loc["15", "min_required"])

    # Pools (normalized names only)
    smt_pool   = df_team.loc[df_team["trained_social"], "norm_name"].tolist()
    all_people = df_team["norm_name"].tolist()
    print("SMT pool size:", len(smt_pool))

    # 1) Pick SMTs
    late_late_team = select_fairest(
        candidates=smt_pool,
        df_history=df_history,
        df_team=df_team,
        shift_type="late_late",
        k=smt_req_15,
        week_start=week_start,
        cooldown_days=6,
        duplicate_teams=False
    )

    remaining_smt_pool = [n for n in smt_pool if n not in late_late_team]
    early_late_team = select_fairest(
        candidates=remaining_smt_pool,
        df_history=df_history,
        df_team=df_team,
        shift_type="early_late",
        k=smt_req_13,
        week_start=week_start,
        cooldown_days=6,
        duplicate_teams=False
    )

    # 2) Fill remaining headcount with non-SMTs
    need_15 = max(0, headcount_15 - len(late_late_team))
    if need_15:
        filler_pool = [n for n in all_people if n not in set(late_late_team) | set(early_late_team)]
        more_15 = select_fairest(
            candidates=filler_pool,
            df_history=df_history,
            df_team=df_team,
            shift_type="late_late",
            k=need_15,
            week_start=week_start,
            cooldown_days=6,
            duplicate_teams=False
        )
        late_late_team += more_15

    need_13 = max(0, headcount_13 - len(early_late_team))
    if need_13:
        filler_pool = [n for n in all_people if n not in set(late_late_team) | set(early_late_team)]
        more_13 = select_fairest(
            candidates=filler_pool,
            df_history=df_history,
            df_team=df_team,
            shift_type="early_late",
            k=need_13,
            week_start=week_start,
            cooldown_days=6,
            duplicate_teams=False
        )
        early_late_team += more_13

    # 3) Write assignments for Mon–Fri
    for nm in late_late_team:
        days = df_week[(df_week["Employee"] == nm) & (df_week["Weekday"] < 5)]["Day"].tolist()
        for d in days:
            assign_one_agent(df_week, d, "15", nm)
    if "assigned_weekly_15" not in df_week.columns:
        df_week["assigned_weekly_15"] = False
    df_week.loc[df_week["Employee"].isin(late_late_team), "assigned_weekly_15"] = True

    for nm in early_late_team:
        days = df_week[(df_week["Employee"] == nm) & (df_week["Weekday"] < 5)]["Day"].tolist()
        for d in days:
            assign_one_agent(df_week, d, "13", nm)
    if "assigned_weekly_13" not in df_week.columns:
        df_week["assigned_weekly_13"] = False
    df_week.loc[df_week["Employee"].isin(early_late_team), "assigned_weekly_13"] = True

    return df_week, late_late_team, early_late_team


def enforce_minimum_staffing(df_schedule, shift_requirements):
    """
    Ensures at least shift_requirements.loc[shift, 'min_required'] people are assigned
    to each (Day, shift_time) in df_schedule.
    """
    # 1. Count current assignments per day & shift
    counts = (
        df_schedule
        .groupby(["Day", "shift_time"])
        .size()
        .rename("assigned_count")
        .reset_index()
    )

    # 2. Merge in the min_required for each shift_time
    counts = counts.merge(
        shift_requirements["min_required"].rename("min_required"),
        left_on="shift_time",
        right_index=True,
        how="left"
    )
    counts = counts[counts["min_required"].notna()]

    # 3. For any (Day, shift_time) where assigned_count < min_required, log or fix
    for _, row in counts.iterrows():
        day, shift_time = row["Day"], row["shift_time"]
        needed = int(row["min_required"] - row["assigned_count"])
        if needed <= 0:
            continue

        print(f"⚠ {day.date()} {shift_time}: need {needed} more staff")

        for _ in range(needed):
            mask_day = df_schedule["Day"] == day
            mask_shift = df_schedule["shift_time"].isin(["9", "RDO"])
            free_df = df_schedule[mask_day & mask_shift]

            if classify_shift(shift_time) == "early_late":
                free_df = free_df[free_df["assigned_weekly_13"] == True]
            elif classify_shift(shift_time) == "late_late":
                free_df = free_df[free_df["assigned_weekly_15"] == True]

            candidates = free_df["Employee"].unique().tolist()
            if not candidates:
                print(f"  ❌ No valid weekly candidate for {day.date()} {shift_time}")
                break

            selected = pick_random_staff(candidates, exclude=None)
            if not selected:
                print(f"  ❌ No available candidate for {day.date()} {shift_time}")
                break

            assign_one_agent(df_schedule, day, shift_time, selected)
            print(f"assigning {selected} to {shift_time} shift on {day.day_name()}")

    return df_schedule

def assign_needed_smt(df_schedule, day, shift, smt_needed, filter_func = None):
    # Filter SMT candidates who are not already assigned to this shift
    available_smt = df_schedule[
        (df_schedule["trained_social"] == True) &
        (df_schedule["Day"] == day) &
        (df_schedule["shift_time"].isna())
    ]

    smt_language = df_schedule[
        (df_schedule["trained_social"] == True) &
        (df_schedule["Day"] == day) &
        (df_schedule["shift_time"] == shift)
    ]["team"].tolist()

    if available_smt.empty:
        print(f"❌ No SMT agents available for shift {shift} on {day}.")
        return df_schedule

    if len(available_smt) < smt_needed:
        print(f"⚠️ Not enough SMT agents for shift {shift} on {day}. Needed: {smt_needed}, Available: {len(available_smt)}")
        return df_schedule

    # Sample SMTs and check language diversity
    selected_smt = available_smt.sample(n=smt_needed)
    selected_lang = selected_smt["team"].tolist()

    if not any(lang in smt_language for lang in selected_lang):
        for _, smt_row in selected_smt.iterrows():
            assign_one_agent(df_schedule, day, shift, smt_row["Employee"])
        print(f"✅ Assigned SMTs to shift {shift} on {day}:", selected_smt["Employee"].tolist())
    else:
        print(f"⚠️ Language overlap detected — skipping SMT assignment for shift {shift} on {day}.")
    selected_lang
    return df_schedule

def assign_one_agent(df_schedule, day, shift, agent):
    try:
        if isinstance(day, str):  # If passed "Monday", pick that day by weekday name
            mask = (
                (df_schedule["Employee"] == agent) &
                (df_schedule["Day"].dt.day_name() == day)
            )
        else:  # If passed a datetime/date, match directly
            mask = (
                (df_schedule["Employee"] == agent) &
                (df_schedule["Day"] == pd.to_datetime(day))
            )
        df_schedule.loc[mask, "shift_time"] = shift
    except Exception as e:
        print(f"⚠️ Error assigning agent {agent} to {day} shift {shift}: {e}")

def classify_shift(shift_start, date=None):
    try:
        hour = int(str(shift_start)[:2])
    except (ValueError, TypeError):
        return None

    if hour < 11:
        return "morning"
    elif hour < 15:
        return "early_late"
    else:
        return "late_late"

def assign_saturday_shifts(df_week, df_history, df_team, grouped_smt_team, shift_requirements, top_n_saturday=3):
    print("✅ START assign_saturday_shifts")
    week = df_week["Day"].dt.isocalendar().week.iloc[0]
    used_languages_late = set()

    already_lated = df_week.loc[
        (df_week["shift_time"].isin(["13", "15"])) &
        (df_week["Day"].dt.day_name() != "Saturday")
    ]["Employee"].unique().tolist()

    already_lated = set(already_lated)
    saturday_eligible = df_team.loc[
        ~df_team["norm_name"].isin(already_lated),
        "name"
    ].tolist()
    print("📊 All SMT candidates available:")
    print(df_team[df_team["trained_social"] == True][["name", "team"]])

    sat_late_late_team = []
    sat_early_late_team = []
    requirement_15 = shift_requirements.loc["15", "min_required"]
    requirement_13 = shift_requirements.loc["13", "min_required"]

    # 1. Assign Saturday 15:00 (late-late)
    while len(sat_late_late_team) < requirement_15:
        smt_requirement = shift_requirements.loc["15", "smt_needed"]

        raw_smt = df_team[df_team["trained_social"] == True]["name"].tolist()
        raw_smt = _norm_name_list(df_team.loc[df_team["trained_social"], "name"].tolist())

        print(df_history[df_history["name"].isin(raw_smt)][["name", "count_saturday_late_late"]])

        print("👀 SMTs available BEFORE fairness check:", raw_smt)
        print("🧾 Counts in df_history:")
        print(df_history[df_history["name"].isin(raw_smt)][["name", "count_saturday_late_late"]])
        fair_sat15 = select_fairest(
            df_history, df_team, saturday_eligible,
            shift_type="saturday_late_late",
            k=k,
            filter_func=lambda row: row["team"] not in used_languages_late and row["trained_social"],
            duplicate_teams=False,
            cooldown_days=6
        )

        print("✅ SMTs selected for fairness (Sat 15):", fair_sat15["name"].tolist())

        if fair_sat15.empty:
            print("❌ No eligible candidate for Saturday 15:00 (late-late)")
            print("🔁 RETURNING EMPTY LISTS")
            return df_week, [], []
        name = fair_sat15["name"].iloc[0]
        lang = df_team.loc[df_team["norm_name"] == name, "team"].iloc[0]
        used_languages_late.add(lang)
        sat_late_late_team.append(name)

    # 2a. Assign one SMT for 13:00 (early-late)
    if requirement_13 > 0:
        smt_requirement = shift_requirements.loc["13", "smt_needed"]
        raw_smt = df_team[df_team["trained_social"] == True]["name"].tolist()
        print("👀 SMTs available BEFORE fairness check:", raw_smt)
        print("🧾 Counts in df_history:")
        print(df_history[df_history["name"].isin(raw_smt)][["name", "count_saturday_late_late"]])

        fair_smt = select_fairest(
            df_history, df_team, saturday_eligible,
            shift_type="saturday_early_late",
            k=k,
            filter_func=lambda row: row["team"] not in used_languages_late and row["trained_social"],
            duplicate_teams=False,
            cooldown_days=6
        )
        print("✅ SMTs selected for fairness (Sat 13):", fair_smt["name"].tolist())

        if fair_smt.empty:
            print("❌ No SMT candidate for Saturday 13:00 (early-late)")
            print("🔁 RETURNING EMPTY LISTS")
            return df_week, [], []
        smt_name = fair_smt["name"].iloc[0]
        lang = df_team.loc[df_team["norm_name"] == smt_name, "team"].iloc[0]
        used_languages_late.add(lang)
        sat_early_late_team.append(smt_name)
    print("📉 SMTs after filtering for 15:00:", fair_sat15["name"].tolist())
    print("📉 SMTs after filtering for 13:00:", fair_smt["name"].tolist())

    # 2b. Fill remaining 13:00 slots
    while len(sat_early_late_team) < requirement_13:
        raw_smt = df_team[df_team["trained_social"] == True]["name"].tolist()
        print("👀 SMTs available BEFORE fairness check:", raw_smt)
        print("🧾 Counts in df_history:")
        print(df_history[df_history["name"].isin(raw_smt)][["name", "count_saturday_late_late"]])

        fair_sat13 = select_fairest(
            df_history, df_team, saturday_eligible,
            shift_type="saturday_early_late",
            k=k,
            filter_func=lambda row: row["team"] not in used_languages_late,
            duplicate_teams=False,
            cooldown_days=6
        )

        if fair_sat13.empty:
            print("❌ No non-SMT candidate for Saturday 13:00 (early-late)")
            print("🔁 RETURNING EMPTY LISTS")
            return df_week, [], []
        name = fair_sat13["name"].iloc[0]
        lang = df_team.loc[df_team["norm_name"] == name, "team"].iloc[0]
        used_languages_late.add(lang)
        sat_early_late_team.append(name)

    # 3. Assign shifts
    for name in sat_late_late_team:
        days = df_week[(df_week["Employee"] == name) &
                       (df_week["Day"].dt.day_name() == "Saturday")]["Day"]
        for d in days:
            assign_one_agent(df_week, d, "15", name)

    for name in sat_early_late_team:
        days = df_week[(df_week["Employee"] == name) &
                       (df_week["Day"].dt.day_name() == "Saturday")]["Day"]
        for d in days:
            assign_one_agent(df_week, d, "13", name)

    print("✅ END assign_saturday_shifts")
    print("🔁 RETURNING NORMAL VALUES")
    return df_week, sat_late_late_team, sat_early_late_team

def assign_saturdays_and_rdo(df_week, df_team, shift_requirements, df_history, sat_late_late_team, sat_early_late_team):
    # ────────── Step 1: Assign Saturday morning ("9") SMT ──────────
    sat_morning_team = set()
    languages_morning_team = set()
    morning_shift = "9"
    morning_smt_needed = shift_requirements.loc["9", "smt_needed"]  # number of SMT needed for Saturday morning
    df_week = assign_needed_smt(df_week, "Saturday", morning_shift, morning_smt_needed)

    requirement_mor = shift_requirements.loc["9", "min_required"]
    day_mask = df_week["Day"].dt.day_name() == "Saturday"
    working_mask = df_week["shift_time"].isin(["RDO", None])
    combined_mask = day_mask & working_mask
    morning_rows = df_week[combined_mask]["Employee"]
    candidates_mor = morning_rows.tolist()

    sat_morning_team = select_fairest(
        df_history,
        df_team,
        candidates_mor,
        shift_type="saturday_morning",
        k=k,
        duplicate_teams=False,
        cooldown_days=6
    )

    for agent in sat_morning_team["name"]:
        assign_one_agent(df_week, "Saturday", morning_shift, agent)
        team_row = df_team.loc[df_team["norm_name"] == agent, "team"]
        if not team_row.empty:
            languages_morning_team.add(team_row.values[0])
        else:
            print(f"⚠️ Agent '{agent}' not found in df_team — cannot retrieve team info.")

    # ────────── Step 2: Gather all Saturday‐involved agents ──────────
    all_saturday_agents = set(
        (sat_early_late_team if isinstance(sat_early_late_team, list) else sat_early_late_team["name"].tolist())
        + (sat_late_late_team if isinstance(sat_late_late_team, list) else sat_late_late_team["name"].tolist())
        + ((sat_morning_team["name"].tolist()) if (
                    'sat_morning_team' in locals() and hasattr(sat_morning_team, "columns")) else (
            sat_morning_team if 'sat_morning_team' in locals() else []))
    )

    all_saturday_agents = [str(n).strip().replace("*", "").lower() for n in all_saturday_agents]

    possible_day_off = ["Wednesday", "Thursday", "Friday"]

    # ────────── Step 3: For each Saturday agent, assign one mid‐week RDO and fill missing Mon–Fri with "9" ──────────
    for i, agent in enumerate(all_saturday_agents):
        # 3a. Pick a rotating weekday to give as RDO
        day_off = possible_day_off[i % len(possible_day_off)]
        df_week.loc[
            (df_week["Employee"] == agent) &
            (df_week["Day"].dt.day_name() == day_off),
            "shift_time"
        ] = "RDO"

        # 3b. Check Mon–Fri actual vs. expected workdays, fill missing with "9"
        mon_to_fri = df_week[
            (df_week["Employee"].str.strip() == agent.strip()) &
            (df_week["Day"].dt.weekday < 5)
        ]

        actual_workdays = mon_to_fri[
            ~mon_to_fri["shift_time"].isin(["RDO", "OFF", "AL", None])
        ].shape[0]

        # Safely look up days_per_week (strip whitespace first)
        agent_clean = agent.strip()
        team_row = df_team[df_team["norm_name"].str.strip() == agent_clean]

        if team_row.empty:
            print(f"⚠️ Agent '{agent}' not found in df_team; skipping their days_per_week lookup")
            continue

        expected_workdays = int(team_row["days_per_week"].values[0])

        if actual_workdays < expected_workdays:
            needed = expected_workdays - actual_workdays
            missing_idx = mon_to_fri[mon_to_fri["shift_time"].isna()].index[:needed]
            df_week.loc[missing_idx, "shift_time"] = "9"

    # ────────── Step 4: Debug print any still‐missing Mon–Fri slots (should be none) ──────────
    missing_workdays = df_week[
        (df_week["Day"].dt.day_name().isin(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])) &
        (df_week["shift_time"].isna())
    ]
    print("\n📅 Missing Mon–Fri shift assignments before fallback to '9':")
    print(missing_workdays[["Employee", "Day", "shift_time"]])

    # ────────── Step 5: Finally set any leftover Saturday slots to "RDO" ──────────
    df_week.loc[
        (df_week["Day"].dt.day_name() == "Saturday") &
        (df_week["shift_time"].isna()),
        "shift_time"
    ] = "RDO"

    print("\n🧮 Days worked just before enforce_max_days_off():")
    print_avg_days_worked(df_week)

    return df_week, sat_morning_team, sat_late_late_team, sat_early_late_team

# In enforce_max_days_off()
def enforce_max_days_off(df_schedule, df_team):
    for _, row in df_team.iterrows():
        agent = row["name"]
        max_rdo_per_week = 7 - row["days_per_week"]
        agent_schedule = df_schedule[df_schedule["Employee"] == agent].copy()
        agent_schedule["temp_week"] = agent_schedule["Day"].dt.isocalendar().week

        for week, week_data in agent_schedule.groupby("temp_week"):
            # Count how many days this agent actually worked (excluding RDO/OFF/AL/None)
            actual_days_worked = week_data[
                ~week_data["shift_time"].isin(["RDO", "OFF", "AL", None])
            ].shape[0]
            expected_days = row["days_per_week"]

            # If already met or exceeded expected days, skip
            if actual_days_worked >= expected_days:
                continue

            rdo_count = week_data[week_data["shift_time"] == "RDO"].shape[0]
            rdo_balance = max_rdo_per_week - rdo_count

            if rdo_balance < 0:
                # Too many RDOs—convert excess RDOs back to "9"
                rdo_rows = week_data[week_data["shift_time"] == "RDO"]
                rows_to_change = rdo_rows.sample(n=abs(rdo_balance)).index
                for idx in rows_to_change:
                    target_day = df_schedule.loc[idx, "Day"]
                    assign_one_agent(df_schedule, target_day.day_name(), "9", agent)

            elif rdo_balance > 0:
                # We need to add RDOs. First look for Saturday shifts to convert.
                saturday_work = week_data[
                    (week_data["shift_time"] != "RDO") &
                    (week_data["Day"].dt.day_name() == "Saturday")
                ]

                if len(saturday_work) < rdo_balance:
                    # Not enough Saturdays to convert, include other working days
                    extra_work = week_data[
                        ~week_data["shift_time"].isin(["RDO", "OFF", "AL"]) &
                        (week_data["Day"].dt.day_name() != "Saturday")
                    ]
                    saturday_work = pd.concat([saturday_work, extra_work])

                # Now filter that combined set to Mon–Fri only
                weekday_candidates = saturday_work[saturday_work["Day"].dt.weekday < 5]

                if len(weekday_candidates) >= rdo_balance:
                    rows_to_change = weekday_candidates.sample(n=rdo_balance).index
                    for idx in rows_to_change:
                        target_day = df_schedule.loc[idx, "Day"]
                        assign_one_agent(df_schedule, target_day.day_name(), "RDO", agent)
                else:
                    print(f"⚠️ Not enough working rows to assign RDOs for agent {agent}")

    return df_schedule

def smt_check(df_schedule, df_team):
    team_shifts = ["9", "13", "15"]
    smt_team = df_team[df_team["trained_social"] == True]["name"].tolist()
    grouped = df_schedule.groupby(["Day", "shift_time"])
    smts_present = []
    for (day, shift), group_df in grouped:
        if shift not in team_shifts:
            continue
        scheduled_agents = group_df["Employee"].tolist()
        smts_present = [agent for agent in scheduled_agents if agent in smt_team]

    if len(smts_present) >= 2:
        print(f"✅ Enough SMT on {day} ({shift}): {len(smts_present)}")
    else:
        print(f"❌ Not enough SMT on {day} ({shift}): {len(smts_present)}")

def report_smt_coverage(df_schedule, df_team):

    smt_team = df_team[df_team["trained_social"] == True]["name"].tolist()
    report_df = df_schedule[df_schedule["Employee"].isin(smt_team)]
    pivot = report_df.pivot_table(index = "shift_time", columns = "Day", values="Employee", aggfunc = "count", fill_value = 0)
    print("SMT report created!")
    return pivot

def _to_name_list(x):
    """Return a list[str] of names no matter if x is list/Series/DataFrame/single."""
    if x is None:
        return []
    if isinstance(x, pd.DataFrame):
        if x.empty:
            return []
        col = "name" if "name" in x.columns else x.columns[0]
        return x[col].dropna().astype(str).tolist()
    if hasattr(x, "dropna"):  # pandas Series
        s = x.dropna()
        return s.astype(str).tolist()
    if isinstance(x, (list, tuple, set)):
        return [str(v) for v in x if v is not None]
    return [str(x)]

def check_and_fill_smt(df_schedule, shift_requirements, df_team):
    smt_needed = {
        "9": 2,
        "15": 1
    }

    smt_agents = df_team[df_team["trained_social"] == True]["name"].tolist()

    for day in df_schedule["Day"].unique():
        if pd.to_datetime(day).weekday() == 6:
            print(f"⛔ Skipping SMT assignment on Sunday: {day}")
            continue

        for shift in ["9", "15"]:
            day_shift_mask = (df_schedule["Day"] == day) & (df_schedule["shift_time"] == shift)
            assigned_smt_count = df_schedule.loc[day_shift_mask & df_schedule["trained_social"], "Employee"].nunique()
            needed = smt_needed[shift] - assigned_smt_count

            if needed > 0:
                pool = df_schedule[
                    (df_schedule["Day"] == day) &
                    (df_schedule["Employee"].isin(smt_agents)) &
                    (df_schedule["shift_time"].isna())
                ]

                print(f"⚠ {day} {shift}: Need {needed} more SMT staff.")

                for agent in pool["Employee"].unique()[:needed]:
                    assign_one_agent(df_schedule, day, shift, agent)
                    print(f"{agent} added to {day} at {shift}")

    return df_schedule

def apply_weekly_shift_logic(df_schedule, df_team, grouped_smt_team, shift_requirements, df_history):
    # tag weeks
    df_schedule["week_id"] = df_schedule["Day"].dt.strftime("%G-W%V")
    all_weeks = sorted(df_schedule["week_id"].unique())

    for week in all_weeks:
        print(f"\n🗖 Assigning shifts for week {week}")
        weekly_mask = df_schedule["week_id"] == week
        df_week = df_schedule[weekly_mask].copy()
        df_week["Weekday"] = df_week["Day"].dt.weekday

        # 1) Assign weekday late-late & early-late
        df_week, late_late_team, early_late_team = assign_both_late_shifts(
            df_week, df_history, df_team, grouped_smt_team, shift_requirements
        )

        # ---------- SAFE helpers ----------
        def _to_list_of_names(obj):
            if obj is None:
                return []
            # DataFrame with a 'name' column
            if hasattr(obj, "columns") and "name" in getattr(obj, "columns", []):
                return obj["name"].astype(str).tolist()
            # already a list/tuple/set of names
            if isinstance(obj, (list, tuple, set)):
                return [str(x) for x in obj]
            # nothing usable
            return []

        # ---------- Use your ACTUAL variables ----------
        # 15:00 = late_late; 13:00 = early_late; 09:00 = morning
        sat15_list = _norm_name_list(_to_list_of_names(locals().get("sat_late_late_team")))
        sat13_list = _norm_name_list(_to_list_of_names(locals().get("sat_early_late_team")))
        sat9_list = _norm_name_list(_to_list_of_names(locals().get("sat_morning_team")))

        updates = {}
        if sat15_list: updates["saturday_late_late"] = sat15_list
        if sat13_list: updates["saturday_early_late"] = sat13_list
        if sat9_list:  updates["saturday_morning"] = sat9_list

        print("Saturday updates summary:", {k: len(v) for k, v in updates.items() if k.startswith("saturday_")})

        if updates:
            picked = set(sum(updates.values(), []))
            cols = [
                "name",
                "count_saturday_late_late", "latest_saturday_late_late",
                "count_saturday_early_late", "latest_saturday_early_late",
                "count_saturday_morning", "latest_saturday_morning",
            ]
            before = df_history.loc[df_history["name"].isin(picked), cols].copy()

            df_history = update_history(df_history, df_week, **updates)

            after = df_history.loc[df_history["name"].isin(picked), cols]
            print("Δ Saturday update (picked only):")
            print(after.merge(before, on="name", suffixes=("_after", "_before")))

        print(df_history.loc[df_history["name"].isin(_to_list_of_names(late_late_team)),
        ["name", "count_late_late", "latest_late_late"]].head())

        print("📈 after weekday update\n",
              df_history[["name", "count_late_late", "count_early_late"]]
              .sort_values(["count_late_late", "count_early_late"], ascending=False)
              .head(10))

        # 2) Assign Saturday SMT (15 & 13)
        result = assign_saturday_shifts(
            df_week, df_history, df_team, grouped_smt_team, shift_requirements, top_n_saturday=3
        )
        if result is None:
            raise RuntimeError("❌ assign_saturday_shifts() returned None unexpectedly!")
        df_week, sat_late_late_team, sat_early_late_team = result

        # Optional: also assign Saturday morning + RDOs (depends on 15/13 picks)
        df_week, sat_morning_team, sat_late_late_team, sat_early_late_team = assign_saturdays_and_rdo(
            df_week,
            df_team,
            shift_requirements,
            df_history,              # use the freshest history
            sat_late_late_team,
            sat_early_late_team
        )
        sat9_list = _to_name_list(sat_morning_team)
        sat13_list = _to_name_list(sat_early_late_team)
        sat15_list = _to_name_list(sat_late_late_team)

        updates = {}
        if sat15_list: updates["saturday_late_late"] = sat15_list
        if sat13_list: updates["saturday_early_late"] = sat13_list
        if sat9_list:  updates["saturday_morning"] = sat9_list

        if updates:
            # sanity: before→after diff for the exact picks
            picked = set(sum(updates.values(), []))
            cols = ["name",
                    "count_saturday_late_late", "latest_saturday_late_late",
                    "count_saturday_early_late", "latest_saturday_early_late",
                    "count_saturday_morning", "latest_saturday_morning"]
            before = df_history.loc[df_history["name"].isin(picked), cols].copy()

            df_history = update_history(df_history, df_week, **updates)

            after = df_history.loc[df_history["name"].isin(picked), cols]
            print("Δ Saturday update (picked only):")
            print(after.merge(before, on="name", suffixes=("_after", "_before")))

        print("📈 after saturday update\n",
              df_history[
                  ["name", "count_saturday_late_late", "count_saturday_early_late", "count_saturday_morning"]
              ]
              .sort_values(["count_saturday_late_late", "count_saturday_early_late", "count_saturday_morning"],
                           ascending=False)
              .head(10))

        # 3) Apply staffing rules (these should not reassign SMT late shifts)
        df_week = enforce_minimum_staffing(df_week, shift_requirements)
        df_week = enforce_max_days_off(df_week, df_team)
        df_week = check_and_fill_smt(df_week, shift_requirements, df_team)

        # 4) Write back this week's assignments into the master schedule ONCE
        df_schedule.loc[weekly_mask, :] = df_week

        # Quick visibility (fix quotes)
        cols = ["name", "count_late_late", "latest_late_late"]
        print(df_history[["name", "count_late_late", "latest_late_late"]]
              .sort_values(["count_late_late", "latest_late_late"])
              .head(10))

    print("\n📅 Weekly shift logic complete for all weeks.")
    return df_history, df_schedule

def output_schedule(excel_file_path, destination_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    # 🗓️ Step 1: Get next and previous months
    try:
        # Extract all sheets that look like "Month YYYY"
        schedule_tabs = []
        for name in wb.sheetnames:
            try:
                parsed = datetime.strptime(name, "%B %Y")
                schedule_tabs.append(parsed)
            except ValueError:
                continue  # Skip tabs that don't match

        if not schedule_tabs:
            raise ValueError("No valid schedule tabs found")

        # Get the most recent tab
        prev_month = max(schedule_tabs)

        # Compute next month manually
        if prev_month.month == 12:
            next_month = datetime(prev_month.year + 1, 1, 1)
        else:
            next_month = datetime(prev_month.year, prev_month.month + 1, 1)

    except Exception as e:
        print(f"Error {e}: no previous schedule found, defaulting to current month")
        today = datetime.today()
        if today.month == 12:
            next_month = datetime(today.year + 1, 1, 1)
            prev_month = datetime(today.year, 12, 1)
        else:
            next_month = datetime(today.year, today.month + 1, 1)
            prev_month = datetime(today.year, today.month, 1)

    next_sheet_name = next_month.strftime("%B %Y")
    prev_sheet_name = prev_month.strftime("%B %Y")
    base_sheet_name = "Base schedule"
    #define next year and month
    year = next_month.year
    month = next_month.month

    # 📁 Step 2: Load workbook and create new sheet from base
    # 📁 Step 2: Load workbook and create new sheet from base
    if base_sheet_name not in wb.sheetnames:
        raise RuntimeError(f"❌ Base sheet '{base_sheet_name}' not found! Sheets found: {wb.sheetnames}")

    try:
        base_sheet = wb[base_sheet_name]
        new_sheet = wb.copy_worksheet(base_sheet)
        new_sheet.title = next_sheet_name
        print(f"📄 Sheetnames before saving: {wb.sheetnames}")
        print(f"✅ Copied new tab: {new_sheet.title}")
    except Exception as e:
        raise RuntimeError(f"❌ Failed to copy base sheet: {e}")

    print("✅ Available sheets:", wb.sheetnames)

    # 📅 Step 3: Determine start and end date
    try:
      previous_sheet = wb[prev_sheet_name]
      for col in reversed(range(3, previous_sheet.max_column + 1)):
        val = previous_sheet.cell(row=1, column=col).value
        if isinstance(val, (datetime, date)):
          previous_sheet_last_date = val
          break
    except KeyError as e:
      print(f"Error {e}: no previous month tab")
      first_day = datetime(year, month, 1)
      previous_sheet_last_date = first_day - timedelta(days=1)


    start_date = previous_sheet_last_date + timedelta(days=1)
    end_of_month = datetime(year, month, calendar.monthrange(year, month)[1]).date()
    end_date = end_of_month + timedelta(days=(6 - end_of_month.weekday()))

    # 📆 Step 4: Write full range of dates into row 1 (starting column 3)
    for i, current_date in enumerate(pd.date_range(start=start_date, end=end_date)):
        new_sheet.cell(row=1, column=3 + i).value = current_date.date()

    # 📋 Step 5: Fill weekday names in each "team" row
    for row in range(1, new_sheet.max_row + 1):
        team_cell = new_sheet.cell(row=row, column=1)
        if team_cell.value and "team" in str(team_cell.value).strip().lower():
            print(f"Found team row: {row} → {team_cell.value}")
            for col in range(3, new_sheet.max_column + 1):
                date_cell = new_sheet.cell(row=1, column=col).value
                if isinstance(date_cell, (datetime, date)):
                    weekday = date_cell.strftime("%A")
                    new_sheet.cell(row=row, column=col).value = weekday
                else:
                    print(f"Warning: cell at row=1, col={col} is not a valid date.")

    #step 6: Fill the formulas at the bottom of the sheet
    latest_row = new_sheet.max_row

    for col in range(3, new_sheet.max_column + 1):
        for row in range(latest_row - 3, latest_row):
            col_letter = get_column_letter(col)
            table_end = latest_row - 5
            new_sheet.cell(row=row, column=col).value = f"=COUNTIF({col_letter}3:{col_letter}{table_end},B{row})"

        new_sheet.cell(row=latest_row, column=col).value = (f"={col_letter}{latest_row-3}+"
                                                                     f"{col_letter}{latest_row-2}+"
                                                                     f"{col_letter}{latest_row-1}")
    #STEP 7: Conditional formatting
    rrdo_fill = PatternFill(start_color="FF707071", end_color="FF707071", fill_type="solid")
    rrdo_rule = CellIsRule(operator='equal', formula=['"RRDO"'], fill=rrdo_fill)
    rdo_fill = PatternFill(start_color="FF908e92", end_color="FF908e92", fill_type="solid")
    rdo_rule = CellIsRule(operator='equal', formula=['"RDO"'], fill=rdo_fill)
    morning_fill = PatternFill(start_color="FFf9f7fb", end_color="FFf9f7fb", fill_type="solid")
    morning_rule = FormulaRule(formula=['LEFT(C3, 2)="09"'], fill=morning_fill)
    early_late_fill = PatternFill(start_color="FFc27aff", end_color="FFc27aff", fill_type="solid")
    early_late_rule = FormulaRule(formula=['LEFT(C3, 2)="13"'], fill=early_late_fill)
    late_late_fill = PatternFill(start_color="FF8903fb", end_color="FF8903fb", fill_type="solid")
    late_late_rule = FormulaRule(formula=['LEFT(C3, 2)="15"'], fill=late_late_fill)
    al_fill = PatternFill(start_color="FF2596be", end_color="FF2596be", fill_type="solid")
    al_rule = CellIsRule(operator='equal', formula=['"AL"'], fill=al_fill)
    off_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    off_font = Font(color="FFFFFF")  # white text

    off_rule = FormulaRule(formula=[f'LOWER(C3)="off"'], fill=off_fill, font=off_font)
    range_str = f"C3:{get_column_letter(new_sheet.max_column)}{new_sheet.max_row}"

    new_sheet.conditional_formatting.add(range_str, rdo_rule)
    new_sheet.conditional_formatting.add(range_str, rrdo_rule)
    new_sheet.conditional_formatting.add(range_str, morning_rule)
    new_sheet.conditional_formatting.add(range_str, early_late_rule)
    new_sheet.conditional_formatting.add(range_str, late_late_rule)
    new_sheet.conditional_formatting.add(range_str, al_rule)
    new_sheet.conditional_formatting.add(range_str, off_rule)


    # 💾 Step 8: Save updated file
    wb.save(destination_file_path)

    # 🧪 Reload and verify the sheet was actually created
    wb = openpyxl.load_workbook(destination_file_path)
    if next_sheet_name not in wb.sheetnames:
        raise RuntimeError(f"❌ Tab '{next_sheet_name}' was not saved correctly in {destination_file_path}")

    # ✅ Return metadata
    sheet_name = next_sheet_name
    return sheet_name, start_date, end_date, pd.date_range(start=start_date, end=end_date)

import sys

def dbg(*args):
    print(*args, flush=True, file=sys.stdout)

def load_employee_history(file_path):

    from dateutil.parser import parse

    print("📥 load_employee_history: starting with", file_path)
    unrecognized_cells = []

    # ─── 1) Try to load the History sheet ────────────────────────────────
    try:
        df = pd.read_excel(file_path, sheet_name="History", usecols="B:L")
        print("✅ History sheet read, shape:", df.shape)
    except Exception as e:
        print(f"⚠️ Cannot read History sheet — {e}")
        print("   falling back to Team sheet for history initialization")
        try:
            df = pd.read_excel(file_path, sheet_name="Team", usecols="B:L")
            print("✅ Team sheet read instead, shape:", df.shape)
            # drop entirely blank rows
            df = df.dropna(how="all")
        except Exception as e2:
            dbg(f"❌ Cannot read Team sheet either — {e2}")
            raise RuntimeError("No History or Team sheet found") from e2

    # ─── 2) Clean & normalize df_history ─────────────────────────────────
    df.columns = df.columns.str.strip().str.lower()
    # rename agent's name → name if present
    if "agent's name" in df.columns:
        df = df.rename(columns={"agent's name": "name"})
        dbg("   renamed \"agent's name\" → \"name\"")
    else:
        dbg("   warning: no \"agent's name\" column to rename")

    # ensure all the count_... cols exist as ints
    zero_cols = [
        "count_early_late", "count_late_late",
        "count_saturday_morning", "count_saturday_early_late", "count_saturday_late_late"
    ]
    for c in zero_cols:
        df[c] = df.get(c, 0).fillna(0).astype(int)
    dbg("   ensured zero-count columns")

    # ensure all the latest_... cols exist as object
    latest_cols = [
        "latest_early_late", "latest_late_late",
        "latest_saturday_morning", "latest_saturday_early_late", "latest_saturday_late_late"
    ]
    for c in latest_cols:
        df[c] = df.get(c, None)
    dbg("   ensured latest-date columns")

    # fill days_per_week default if missing
    if "days_per_week" not in df.columns or df["days_per_week"].isnull().any():
        if "days_per_week" not in df.columns:
            df["days_per_week"] = 5
        else:
            df["days_per_week"] = df["days_per_week"].fillna(5).astype(int)

        dbg("   filled days_per_week default = 5")

    # strip & lower the names so matching always works
    if "name" in df.columns:
        df["name"] = df["name"].astype(str).str.strip().str.lower()
        dbg("   normalized names")
    else:
        dbg("   warning: no \"name\" column found after renaming")

    # ─── 3) Grab the latest schedule sheet to return ──────────────────────
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_dates = []
    for name in wb.sheetnames:
        try:
            dt = parse(name)
            sheet_dates.append((name, dt))
        except:
            pass
    if not sheet_dates:
        raise RuntimeError("No date-formatted sheets found to pick latest schedule from")
    sheet_dates.sort(key=lambda x: x[1])
    latest_name = sheet_dates[-1][0]
    schedule_sheet = wb[latest_name]
    print(f"✅ Using latest schedule tab: \"{latest_name}\"")

    # ─── 4) Done! ────────────────────────────────────────────────────────
    print(f"   🎯 load_employee_history complete, returning df({df.shape}), "
          f"cells({len(unrecognized_cells)}), sheet({latest_name})")
    return df, unrecognized_cells, schedule_sheet

def check_mismatches(excel_file_path, df_history, schedule_sheet, df_team):
    """
    Compares names across different tabs:
    - Team tab
    - History tab
    - Latest schedule sheet
    - Holiday sheet (optional)

    Returns a dictionary of mismatches.
    """
    previous_schedule_sheet = schedule_sheet  # to be returned from the load employee function
    base_schedule_sheet = pd.read_excel(excel_file_path, sheet_name="Base schedule", usecols="A:B")
    from openpyxl.utils import get_column_letter

    # Start with openpyxl
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)

    if "Holidays" not in wb.sheetnames:
        print("❌ 'Holidays' sheet not found in workbook.")
        holiday_sheet = pd.DataFrame()  # or handle differently if needed
    else:
        ws = wb["Holidays"]

        # Find last used column in row 1
        last_used_col = 2
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=1, column=col).value:
                last_used_col = col

        from openpyxl.utils import get_column_letter
        end_col_letter = get_column_letter(last_used_col)
        usecols = f"B:{end_col_letter}"
        print(f"✅ Reading Holidays with usecols='{usecols}'")

        holiday_sheet = pd.read_excel(excel_file_path, sheet_name="Holidays", usecols=usecols)

    team_names = [str(name).strip().lower() for name in df_team["norm_name"].dropna()]
    previous_schedule_names = []
    history_names = [str(name).strip().lower() for name in df_history["name"].dropna()]
    holiday_names = []
    base_schedule_names = []

    # Finding the previous schedule names
    for row in previous_schedule_sheet.iter_rows(min_col=2, max_col=2):
        for cell in row:
            name = cell.value
            if name:
                previous_schedule_names.append(name.strip().lower())

    # Finding the holiday names (from row 1)
    holiday_names = [
        str(val).strip().lower()
        for val in holiday_sheet.iloc[0, :]
        if pd.notna(val) and str(val).strip()
    ]

    # Finding the base schedule names
    for idx, row in base_schedule_sheet.iterrows():
        for cell in row:
            if cell:
                base_schedule_names.append(str(cell).strip().lower())

    leavers = set(previous_schedule_names) - set(team_names)
    new_hires = set(team_names) - set(previous_schedule_names)
    no_history = set(team_names) - set(history_names)
    history_only = set(history_names) - set(team_names)
    missing_in_team = set(base_schedule_names) - set(team_names)
    missing_in_base_schedule = set(team_names) - set(base_schedule_names)
    missing_in_holiday = set(team_names) - set(holiday_names)
    missing_from_holiday = set(holiday_names) - set(team_names)

    mismatches = {
        "In previous months but not team sheet": leavers,
        "In the team sheet but not in the previous months": new_hires,
        "People in the team sheet not in the history sheet": no_history,
        "People in the history sheet not in the base schedule": history_only,
        "people in the base schedule but not in the team sheet": missing_in_team,
        "people in the team sheet but not in the base schedule": missing_in_base_schedule,
        "people in the holiday sheet but not in the team sheet": missing_from_holiday,
        "people in the team sheet but not in the holiday sheet": missing_in_holiday,
    }

    return mismatches

def normalize(name):
    return str(name).strip().lower()

def fill_schedule(df_team, df_schedule, destination_file_path, sheet_name, name_map):
    print(f"📂 Filling schedule into: {destination_file_path}")
    print("📋 Sheet name:", sheet_name)
    print("📊 df_schedule shape:", df_schedule.shape)
    print("👥 df_schedule Employees:", df_schedule['Employee'].unique())
    wb = openpyxl.load_workbook(destination_file_path)
    sheet = wb[sheet_name]
    staff_list = df_team["norm_name"].tolist()

    print("👤 Example mapping:")
    for nm in staff_list[:5]:
        print(f"norm={nm!r} → excel={name_map.get(nm)!r}")

    for norm_name in staff_list:
        df_agent = df_schedule[df_schedule["Employee"] == norm_name]
        if df_agent.empty:
            print(f"⚠️ No schedule for: {norm_name}")
            continue

        # translate back to Excel name for row lookup
        excel_name = name_map.get(norm_name)
        if not excel_name:
            print(f"⚠️ No mapping for: {norm_name}")
            continue

        agent_row = df_team[df_team["name"] == excel_name]
        if agent_row.empty:
            print(f"⚠️ No Excel row for: {excel_name}")
            continue

        hours_per_week = agent_row["hours_per_week"].values[0]
        days_per_week = agent_row["days_per_week"].values[0]

        # --- row lookup in column B, starting row 2 ---
        row_to_fill = None
        for col_cells in sheet.iter_cols(min_col=2, max_col=2, min_row=2):
            for cell in col_cells:
                cell_value = str(cell.value).replace("\u00A0", " ").strip() if cell.value else ""
                if cell_value == excel_name or cell_value.rstrip("*").lower() == excel_name.rstrip("*").lower():
                    row_to_fill = cell.row
                    break
            if row_to_fill is not None:
                break

        if row_to_fill is None:
            print(f"⚠️ Could not find row for: {excel_name}")
            continue

        # --- write shifts for this agent (your code continues here) ---
        for col in range(3, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            wb_date = cell.value
            if wb_date is None: break
            if not isinstance(wb_date, (datetime, date)): continue

            wb_day = pd.to_datetime(wb_date).date()
            for _, r in df_agent.iterrows():
                row_day = pd.to_datetime(r["Day"]).date()
                if wb_day == row_day:
                    shift = r["shift_time"]
                    if isinstance(shift, str) and shift.isdigit():
                        start_hour = int(shift)
                        days_used = 5 if classify_shift(shift) in {"early_late", "late_late"} else days_per_week
                        hours_per_day = hours_per_week / days_used
                        end_hour = start_hour + hours_per_day + 1
                        shift_str = f"{start_hour:02.0f}:00 - {int(end_hour):02.0f}:00"
                        sheet.cell(row=row_to_fill, column=col).value = shift_str
                    else:
                        sheet.cell(row=row_to_fill, column=col).value = shift

    print("✅ Writing file to:", destination_file_path)
    print("✅ Schedule preview:\n", df_schedule.head(10))

    wb.save(destination_file_path)

def fill_history_tab(df_history, excel, sheet_name="History", all_months=False):
    print("📥 fill_history_tab called with:", excel)
    try:
        wb = openpyxl.load_workbook(excel, data_only=True)
        print("   ✅ workbook loaded, sheets:", wb.sheetnames)
    except Exception as e:
        print("   ❌ fill_history_tab: cannot open workbook:", e)
        raise

    dbg("   sheets found:", wb.sheetnames)
    # 1) Parse every “Month YYYY” tab
    parsed_tabs = []
    today = datetime.today().replace(day=1)

    for name in wb.sheetnames:
        try:
            parsed_date = datetime.strptime(name, "%B %Y")
            parsed_tabs.append((name, parsed_date))

        except ValueError:
            continue
    valid_tabs = [(n, d) for n, d in parsed_tabs if d < today]

    print("Parsed tabs:", parsed_tabs)
    print("Today is:", today)
    print("Valid tabs:", valid_tabs)

    # 2) Bail out if there were no date-named tabs at all
    if not parsed_tabs:
        raise ValueError("No valid schedule tabs found in the workbook.")

    # 3) Filter out current/future months, keep only past
    if not valid_tabs:
        raise ValueError("No past schedule tabs to process.")
    # 4) Sort the filtered list by date
    valid_tabs.sort(key=lambda x: x[1])

    # 5) Decide which tab(s) to process
    if all_months:
        tabs_to_process = [name for name, _ in valid_tabs]
    else:
        tabs_to_process = [valid_tabs[-1][0]]  # only the most‐recent past month

    # 6) Prepare to update the History sheet
    history_sheet = wb[sheet_name]
    staff_list   = df_history["name"].tolist()

    # Map header names → column indices
    col_indices = {
        history_sheet.cell(row=1, column=col).value: col
        for col in range(1, history_sheet.max_column + 1)
    }
    # Sanity-check that all required headers exist
    required = [
        "latest_early_late", "count_early_late",
        "latest_late_late",  "count_late_late",
        "latest_saturday_morning",  "count_saturday_morning",
        "latest_saturday_early_late", "count_saturday_early_late",
        "latest_saturday_late_late",  "count_saturday_late_late"
    ]
    for header in required:
        assert header in col_indices, f"Missing History header: {header}"

    # 7) Process each selected tab
    dbg("   processing History tab rows …")
    for tab in tabs_to_process:
        print(f"Checking tab: {tab}")
        schedule_sheet = wb[tab]
        print(f"staff list to be processed: {staff_list}")
        for staff in staff_list:
            staff = staff.strip().lower()

            # find the row in History and in Schedule
            row_to_fill = next(
                (r[1].row for r in history_sheet.iter_rows(min_row=2, max_col=2)
                 if isinstance(r[1].value, str) and r[1].value.strip().lower() == staff),
                None
            )
            row_to_pull = next(
                (r[1].row for r in schedule_sheet.iter_rows(min_row=3, max_col=2)
                 if isinstance(r[1].value, str) and r[1].value.strip().lower() == staff),
                None
            )
            print(f"row to pull is {row_to_pull}")
            print(f"row to fill is {row_to_fill}")
            if row_to_fill is None or row_to_pull is None:
                print(f"⚠️ No match found for '{staff}' → skipping.")
                continue
            print(f"Processing staff: {staff} → row_to_fill: {row_to_fill}, row_to_pull: {row_to_pull}")
            # collect shift dates by type
            shifts_13_weekday   = []
            shifts_13_saturday  = []
            shifts_15_weekday   = []
            shifts_15_saturday  = []
            shifts_9_saturday   = []
            for col in range(3, schedule_sheet.max_column + 1):
                shift = schedule_sheet.cell(row=row_to_pull, column=col).value

                shift_start = None

                if isinstance(shift, str):
                    s = shift.strip().upper()

                    # Skip markers that are not shifts
                    if s in {"RDO", "RRDO", "AL", "OFF"} or not s:
                        print(f"⚠️ Non-working marker {s!r}, skipping")
                        continue

                    # Handle formats like "08:00-17:00"
                    if "-" in s and s[0:2].isdigit():
                        shift_start = s[0:2]  # take first two chars → "08", "09", "13", "15"
                    elif s.isdigit():
                        shift_start = s
                elif isinstance(shift, (int, float)):
                    shift_start = str(int(shift))

                if not shift_start:
                    print(f"⚠️ Unrecognized shift value {shift!r}, skipping")
                    continue

                # normalize to plain hour
                if shift_start.startswith("0"):
                    shift_start = shift_start.lstrip("0")

                # accept 8, 9, 13, 15
                if shift_start not in {"8", "9", "13", "15"}:
                    print(f"⚠️ Shift start {shift_start!r} not tracked, skipping")
                    continue

                date_value = schedule_sheet.cell(row=1, column=col).value
                weekday_name = date_value.strftime("%A") if isinstance(date_value, datetime) else "Unknown"
                print(f"\n📅 Evaluating date: {date_value}, Weekday: {weekday_name}")
                print(f"🔎 Raw shift value: {shift!r} → Parsed: {shift_start!r}")

                shift_start = int(shift_start)


                date_value = schedule_sheet.cell(row=1, column=col).value
                weekday_name = date_value.strftime("%A") if isinstance(date_value, datetime) else "Unknown"
                print(f"\n📅 Evaluating date: {date_value}, Weekday: {weekday_name}")
                print(f"🔎 Raw shift value: {shift!r} → Parsed: {shift_start!r}")

                # now shift_start is a clean string like "9", "13", or "15"
                shift_start = int(shift_start)

                # Now classify
                try:
                    category = classify_shift(shift_start)
                    print(f"🧪 Classified shift: {category}")
                except Exception as e:
                    print(f"❌ classify_shift failed for {shift_start} → {e}")
                    continue

                # Append to the correct list
                if category == "early_late":
                    if weekday_name == "Saturday":
                        shifts_13_saturday.append(date_value)
                        print(f"📌 Appended to shifts_13_saturday → {date_value}")
                    else:
                        shifts_13_weekday.append(date_value)
                        print(f"📌 Appended to shifts_13_weekday → {date_value}")
                elif category == "late_late":
                    if weekday_name == "Saturday":
                        shifts_15_saturday.append(date_value)
                        print(f"📌 Appended to shifts_15_saturday → {date_value}")
                    else:
                        shifts_15_weekday.append(date_value)
                        print(f"📌 Appended to shifts_15_weekday → {date_value}")
                elif category == "morning" and weekday_name == "Saturday":
                    shifts_9_saturday.append(date_value)
                    print(f"📌 Appended to shifts_9_saturday → {date_value}")

            # helper to write latest and count once‐per‐week
            def update_column(dates, latest_key, count_key, row_to_fill, col_indices, history_sheet, staff):
                print("updating column")
                print(f"dates are the following: {dates}")

                if not dates:
                    return

                seen_weeks = set()
                final_dates = []
                for dt in dates:
                    wk = dt.isocalendar()[1]
                    if wk not in seen_weeks:
                        seen_weeks.add(wk)
                        final_dates.append(dt)

                print(f"final_dates = {final_dates}")
                if not final_dates:
                    print("⚠️ No valid weekly dates to update — skipping write.")
                    return

                latest_shift = max(final_dates)
                print(f"✅ Updating {staff} → {latest_key}: latest = {latest_shift}, count = {len(final_dates)}")

                print("Writing to sheet: History")
                history_sheet.cell(row=row_to_fill, column=col_indices[latest_key]).value = latest_shift
                cell = history_sheet.cell(row=row_to_fill, column=col_indices[count_key])
                cell.value = (cell.value or 0) + len(final_dates)

            # apply to each shift‐type bucket
            print("launching update_column function")
            update_column(shifts_13_weekday,   "latest_early_late",           "count_early_late", row_to_fill, col_indices, history_sheet, staff)
            update_column(shifts_13_saturday,  "latest_saturday_early_late",  "count_saturday_early_late", row_to_fill, col_indices, history_sheet, staff)
            update_column(shifts_15_weekday,   "latest_late_late",            "count_late_late", row_to_fill, col_indices, history_sheet, staff)
            update_column(shifts_15_saturday,  "latest_saturday_late_late",   "count_saturday_late_late", row_to_fill, col_indices, history_sheet, staff)
            update_column(shifts_9_saturday,   "latest_saturday_morning",     "count_saturday_morning", row_to_fill, col_indices, history_sheet, staff)

    # 8) Save all updates back to the workbook
    print("✅ History updated and saved to", excel)

    # at very end of fill_history_tab
    wb.save(excel)
    print(f"✅ fill_history_tab: saved History into {excel!r}")

def print_avg_days_worked(df_schedule):
    temp = df_schedule.copy()
    temp["Week"] = temp["Day"].dt.isocalendar().week
    worked = temp[~temp["shift_time"].isin(["RDO", "OFF", "AL", None])]
    avg_days = (
        worked.groupby(["Employee", "Week"])
        .size()
        .groupby("Employee")
        .mean()
        .sort_values(ascending=False)
    )
    print("\n📊 Average Days Worked Per Week:")
    print(avg_days.round(2))
import os
import shutil

def run_schedule(uploaded_path, output_folder):
    # 1️⃣ Make a working copy right away
    os.makedirs(output_folder, exist_ok=True)
    dest = os.path.join(output_folder, "Team_updated.xlsx")
    shutil.copy(uploaded_path, dest)
    excel = dest
    print("🔨 Working copy is:", excel)
    # 2️⃣ Now load EVERYTHING from the copy
    shift_requirements = load_shift_requirements(excel, sheet_name="Shifts")

    df_team, name_map = load_employee_list(excel)
    print("name_map sample:", list(name_map.items())[:5])
    # RIGHT side must be the exact Excel strings (with caps, possibly trailing *)

    df_history, unrec_cells, schedule_sheet = load_employee_history(excel)

    # 3️⃣ Update that copy’s History tab
    print("📥 run_schedule: about to call fill_history_tab…")
    fill_history_tab(df_history, excel)
    print("   🔍 Checking History[2,3]…", end=" ")

    wb2 = openpyxl.load_workbook(excel, data_only=True)
    val = wb2["History"].cell(row=2, column=3).value
    print(repr(val))

    # 4️⃣ Re-load history/sheet from the same copy to pick up your writes


    # 5️⃣ Compute mismatches against that now-up-to-date copy
    mismatches = check_mismatches(excel,
                                  df_history,
                                  schedule_sheet,
                                  df_team)
    print("✅ run_schedule: mismatches =", mismatches)

    # 6️⃣ Append a new month tab to the copy
    sheet_name, start_date, end_date, full_date_range = output_schedule(
        excel,  # source for base & prev-month
        excel   # and also target for saving
    )

    # 7️⃣ Build your DataFrame and merge
    df_schedule = create_schedule_structure(
        df_holidays=None,
        employee_list=df_team["name"].tolist(),
        start_date=start_date,
        end_date=end_date
    )
    df_schedule = merge_df_team_and_df_schedule(df_team, df_schedule)
    df_schedule = assign_initial_shifts(df_schedule)

    # 8️⃣ Integrate holidays, assign shifts, etc...
    df_holidays      = load_holiday_calendar(excel, sheet_name="Holidays")
    df_holidays_long = process_holiday_calendar(df_holidays)
    if df_holidays_long is not None:
        df_schedule = integrate_holidays_into_schedule(df_schedule, df_holidays_long)

    _, grouped_smt_team = extract_smt_staff_grouped(excel)
    df_history, df_schedule = apply_weekly_shift_logic(
        df_schedule, df_team, grouped_smt_team, shift_requirements, df_history
    )
    # … plus your other assignment passes …

    print("🧪 df_schedule columns:", df_schedule.columns.tolist())
    print(df_history[["name", "count_saturday_late_late", "count_saturday_early_late"]].sort_values(
        by="count_saturday_late_late", ascending=False))

    # 9️⃣ Finally fill that new sheet in the SAME copy
    fill_schedule(df_team, df_schedule, excel, sheet_name, name_map)

    # 🔟 Return everything
    return sheet_name, start_date, end_date, full_date_range, mismatches, unrec_cells, dest

if __name__ == "__main__":

    sheet_name, start_date, end_date, full_date_range, mismatches, unrecognized_cells, output = run_schedule(
        "/home/julien/Documents/source_file.xlsx",
        "/home/julien/Documents/"
    )
    print(f"✅ File saved to: {output}")
